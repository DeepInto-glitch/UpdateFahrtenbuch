import sys
import logging
import json
import subprocess
import requests
import shutil
import os
from PyQt5.QtWidgets import (QApplication, QWidget, QLabel, QPushButton, 
                             QVBoxLayout, QHBoxLayout, QLineEdit, QComboBox, QListWidget, QMessageBox, 
                             QStackedWidget, QFrame, QTableWidget, QTableWidgetItem, QInputDialog)
from PyQt5.QtCore import Qt
from PyQt5 import QtWidgets
from PyQt5 import QtWidgets as QtW
from PyQt5.QtWidgets import QDateEdit, QDialog
from PyQt5.QtCore import QPoint
from openpyxl import Workbook, load_workbook
from datetime import datetime
from datetime import date
from pathlib import Path

# Basisverzeichnis für die App (das Verzeichnis, in dem das Python-Skript liegt)
base_dir = Path(__file__).parent

# Backup-Verzeichnis
BACKUP_DIR = base_dir / "Backup"

# Dateipfad für die JSON-Dateien und Excel-Datei
DATA_FILE = base_dir / "Data" / "strecken_daten.json"
EXCEL_FILE = base_dir / "Excel" / "fahrtenbuch.xlsx"
PATIENTEN_FILE = base_dir / "Data" / "patienten_daten.json"
LOG_FILE = base_dir / "Data" / "logs.log"
VERSION_FILE = base_dir / "Data" / "version.json"
GITHUB_REPO_URL = 'https://raw.githubusercontent.com/DeepInto-glitch/UpdateFahrtenbuch/refs/heads/main/version.json'
UPDATE_FILE = base_dir / "Data" / "Updater.py"

# Erstelle einen Logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.CRITICAL)

# Erstelle einen FileHandler, der die Log-Meldungen in die Datei schreibt
file_handler = logging.FileHandler(LOG_FILE)
file_handler.setLevel(logging.CRITICAL)

# Erstelle einen Formatter, der die Log-Meldungen formatiert
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)

# Füge den FileHandler zum Logger hinzu
logger.addHandler(file_handler)

# Backup erstellen 
def Create_Backup():
    if not os.path.exists(BACKUP_DIR):
        os.makedirs(BACKUP_DIR)
        logger.info('Backupverzeichnis erstellt')

    current_date = date.today().strftime("%Y-%m-%d")

    # Create a folder for the current date
    current_date_folder = os.path.join(BACKUP_DIR, current_date)
    if not os.path.exists(current_date_folder):
        os.makedirs(current_date_folder)

    # Define the file names with the current date
    backup_data_file = f"strecken_daten.json"
    backup_excel_file = f"fahrtenbuch.xlsx"
    backup_patienten_file = f"patienten_daten.json"

    # Define the backup file paths
    backup_data_file_path = os.path.join(current_date_folder, backup_data_file)
    backup_excel_file_path = os.path.join(current_date_folder, backup_excel_file)
    backup_patienten_file_path = os.path.join(current_date_folder, backup_patienten_file)

    # Check if a backup for the current date already exists
    if os.path.exists(backup_data_file_path) and os.path.exists(backup_excel_file_path) and os.path.exists(backup_patienten_file_path):
        print(f"Backup für den {current_date} bereits vorhanden. Kein neues Backup erstellt!")
        logger.warning('Backup existiert bereits')
        logger.info('Kein neues Backup erstellt')
        return

    # Copy the files to the backup directory
    shutil.copyfile(DATA_FILE, backup_data_file_path)
    shutil.copyfile(EXCEL_FILE, backup_excel_file_path)
    shutil.copyfile(PATIENTEN_FILE, backup_patienten_file_path)

    # Get a list of all backup folders
    backup_folders = os.listdir(BACKUP_DIR)

    # If there are more than 10 backup folders, delete the oldest one
    if len(backup_folders) > 10:
        oldest_folder = min(backup_folders, key=lambda x: os.path.getctime(os.path.join(BACKUP_DIR, x)))
        shutil.rmtree(os.path.join(BACKUP_DIR, oldest_folder))
        logger.info('Ältestes Backup wurde gelöscht')

    print(f"Backup wurde für den {current_date} erstellt!")
    logger.info('Backup wurde erstellt')

# Funktion zum Laden der Streckendaten aus der JSON-Datei
def load_strecken_daten():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
            logger.info('Streckendaten geladen')
    else:
        logger.critical('Datei mit Streckendaten wurde nicht gefunden')
        return []

# Funktion zum Speichern der Streckendaten in die JSON-Datei
def save_strecken_daten(strecken_daten):
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(strecken_daten, f, ensure_ascii=False)
        logger.info('Streckendaten wurden gespeichert')

# Funktion zum Laden der Patientendaten aus der JSON-Datei
def load_patienten_daten():
    if os.path.exists(PATIENTEN_FILE):
        with open(PATIENTEN_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
            logger.info('Patientendaten wurden geladen')
    else:
        logger.critical('Datei mit Patientendaten wurde nicht gefunden')
        return []

# Funktion zum Speichern der Patientendaten in die JSON-Datei
def save_patienten_daten(patienten_daten):
    with open(PATIENTEN_FILE, 'w', encoding='utf-8') as f:
        json.dump(patienten_daten, f, ensure_ascii=False)
        logger.info('Patientendaten wurden gespeichert')

# Initialisiere die Strecken- und Patientendaten aus den JSON-Dateien
strecken_daten = load_strecken_daten()
patienten_daten = load_patienten_daten()

# Backup kopieren

Create_Backup()


class CustomWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setStyleSheet("background-color: #2F3136; border-radius: 10px;")
        self.init_ui()
        self.fahrten = []  # Liste für die Fahrten
        self.oldPos = self.pos()

    def mousePressEvent(self, event):
        self.oldPos = event.globalPos()

    def mouseMoveEvent(self, event):
        delta = QPoint(event.globalPos() - self.oldPos)
        self.move(self.x() + delta.x(), self.y() + delta.y())
        self.oldPos = event.globalPos()

    def init_ui(self):
        self.setGeometry(200, 100, 800, 500)
        self.setStyleSheet("background-color: #2F3136; border-radius: 10px")  # Discord-like theme

        # Hauptlayout
        main_layout = QHBoxLayout(self)

        # Menü erstellen
        menu_frame = QFrame(self)
        menu_frame.setFixedWidth(150)
        menu_frame.setStyleSheet("background-color: #23272A; border-radius: 10px;")

        menu_layout = QVBoxLayout(menu_frame)

        # Menüpunkte erstellen
        self.menu_button_fahrten = QPushButton("Hauptmenü")
        self.menu_button_fahrten.clicked.connect(self.show_fahrten)
        
        self.menu_button_tabelle = QPushButton("Strecken")
        self.menu_button_tabelle.clicked.connect(self.show_tabelle)
        
        self.menu_button_patienten = QPushButton("Patienten")
        self.menu_button_patienten.clicked.connect(self.show_patienten)

        self.menu_button_sonstiges = QPushButton("Exportieren")
        self.menu_button_sonstiges.clicked.connect(self.exportieren)

        self.menu_button_update = QPushButton("Update")
        self.menu_button_update.clicked.connect(self.check_version)
        
        self.menu_button_schliessen = QPushButton("Schließen")
        self.menu_button_schliessen.clicked.connect(self.close_app)

        self.menu_button_fahrten.setStyleSheet("""QPushButton {
            background-color: #2C2F33;
            color: #FFFFFF;
            border-radius: 10px;
            padding: 10px;
            font-size: 16px;
        }
        QPushButton:hover {
            background-color: #7289DA;
        }""")

        self.menu_button_tabelle.setStyleSheet("""QPushButton {
            background-color: #2C2F33;
            color: #FFFFFF;
            border-radius: 10px;
            padding: 10px;
            font-size: 16px;
        }
        QPushButton:hover {
            background-color: #7289DA;
        }""")

        self.menu_button_patienten.setStyleSheet("""QPushButton {
            background-color: #2C2F33;
            color: #FFFFFF;
            border-radius: 10px;
            padding: 10px;
            font-size: 16px;
        }
        QPushButton:hover {
            background-color: #7289DA;
        }""")

        self.menu_button_sonstiges.setStyleSheet("""QPushButton {
            background-color: #2C2F33;
            color: #FFFFFF;
            border-radius: 10px;
            padding: 10px;
            font-size: 16px;
        }
        QPushButton:hover {
            background-color: #7289DA;
        }""")

        self.menu_button_update.setStyleSheet("""QPushButton {
            background-color: #2C2F33;
            color: #FFFFFF;
            border-radius: 10px;
            padding: 10px;
            font-size: 16px;
        }
        QPushButton:hover {
            background-color: #7289DA;
        }""")

        self.menu_button_schliessen.setStyleSheet("""QPushButton {
            background-color: #2C2F33;
            color: #FFFFFF;
            border-radius: 10px;
            padding: 10px;
            font-size: 16px;
        }
        QPushButton:hover {
            background-color: #7289DA;
        }""")

        menu_layout.addWidget(self.menu_button_fahrten)
        menu_layout.addWidget(self.menu_button_tabelle)
        menu_layout.addWidget(self.menu_button_patienten)
        menu_layout.addWidget(self.menu_button_sonstiges)
        menu_layout.addStretch()
        menu_layout.addWidget(self.menu_button_update)
        menu_layout.addWidget(self.menu_button_schliessen)

        # Content-Bereich
        self.stacked_widget = QStackedWidget(self)

        # "Fahrten"-Seite erstellen
        self.fahrten_page = QWidget()
        self.fahrten_layout = QVBoxLayout(self.fahrten_page)

        # Neue ComboBox für Patientenauswahl hinzufügen
        self.patient_label = QLabel("Patient:")
        self.patient_label.setStyleSheet("color: #FFFFFF;")
        self.fahrten_layout.addWidget(self.patient_label)

        self.patient_combobox = QComboBox(self)
        self.patient_combobox.setEditable(True)  # Damit die ComboBox editierbar ist
        self.patient_combobox.setLineEdit(QLineEdit())  # Damit ein QLineEdit-Widget in der ComboBox verwendet wird
        self.update_patient_combobox()  # Patientendaten in die ComboBox laden
        self.patient_combobox.setStyleSheet("""
            QComboBox {
                background-color: #343B41;
                color: #FFFFFF;
                border: 1px solid #7289DA;
                border-radius: 10px;
                padding: 10px;
            }
            QComboBox QAbstractItemView {
                background-color: #343B41;
                color: #FFFFFF;
                border: 1px solid #7289DA;
                border-radius: 10px;
                padding: 10px;
            }
            QComboBox QAbstractItemView::item {
                background-color: #343B41;
                color: #FFFFFF;
                padding: 10px;
            }
            QComboBox QAbstractItemView::item:selected {
                background-color: #7289DA;
                color: #FFFFFF;
            }
            QComboBox QAbstractItemView::item:hover {
                background-color: #7289DA;
                color: #FFFFFF;
            }
        """)

        self.fahrten_layout.addWidget(self.patient_combobox)

        # Datumseingabe
        self.date_label = QLabel("Datum:")
        self.date_label.setStyleSheet("color: #FFFFFF;")
        self.fahrten_layout.addWidget(self.date_label)

        # Datum-Layout
        self.date_layout = QHBoxLayout()
        self.fahrten_layout.addLayout(self.date_layout)

        # Jahr
        self.year_combobox = QComboBox(self)
        self.year_combobox.setEditable(True)  # Damit die ComboBox editierbar ist
        self.year_combobox.setLineEdit(QLineEdit())  # Damit ein QLineEdit-Widget in der ComboBox verwendet wird
        self.year_combobox.addItems([str(year) for year in range(2024, 2040)])  # Jahre von 2024 bis 2040
        self.year_combobox.setCurrentIndex(0)  # Aktuelles Jahr als Standardwert
        self.year_combobox.setStyleSheet("""
            QComboBox {
                background-color: #343B41;
                color: #FFFFFF;
                border: 1px solid #7289DA;
                border-radius: 10px;
                padding: 10px;
            }
            QComboBox QAbstractItemView {
                background-color: #343B41;
                color: #FFFFFF;
                border: 1px solid #7289DA;
                border-radius: 10px;
                padding: 10px;
            }
            QComboBox QAbstractItemView::item {
                background-color: #343B41;
                color: #FFFFFF;
                padding: 10px;
            }
            QComboBox QAbstractItemView::item:selected {
                background-color: #7289DA;
                color: #FFFFFF;
            }
            QComboBox QAbstractItemView::item:hover {
                background-color: #7289DA;
                color: #FFFFFF;
            }
        """)
        self.date_layout.addWidget(self.year_combobox)

        # Monat
        self.month_combobox = QComboBox(self)
        self.month_combobox.setEditable(False)  # Damit die ComboBox nicht editierbar ist
        self.month_combobox.addItems(["Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember"])
        self.month_combobox.setCurrentIndex(datetime.today().month - 1)  # Aktueller Monat als Standardwert
        self.month_combobox.setStyleSheet("""
            QComboBox {
                background-color: #343B41;
                color: #FFFFFF;
                border: 1px solid #7289DA;
                border-radius: 10px;
                padding: 10px;
            }
            QComboBox QAbstractItemView {
                background-color: #343B41;
                color: #FFFFFF;
                border: 1px solid #7289DA;
                border-radius: 10px;
                padding: 10px;
            }
            QComboBox QAbstractItemView::item {
                background-color: #343B41;
                color: #FFFFFF;
                padding: 10px;
            }
            QComboBox QAbstractItemView::item:selected {
                background-color: #7289DA;
                color: #FFFFFF;
            }
            QComboBox QAbstractItemView::item:hover {
                background-color: #7289DA;
                color: #FFFFFF;
            }
        """)
        self.date_layout.addWidget(self.month_combobox)

        # Tag
        self.day_combobox = QComboBox(self)
        self.day_combobox.setEditable(True)  # Damit die ComboBox editierbar ist
        self.day_combobox.setLineEdit(QLineEdit())  # Damit ein QLineEdit-Widget in der ComboBox verwendet wird
        self.day_combobox.addItems([str(day) for day in range(1, 32)])  # Tage von 1 bis 31
        self.day_combobox.setCurrentIndex(datetime.today().day - 1)  # Aktueller Tag als Standardwert
        self.day_combobox.setStyleSheet("""
            QComboBox {
                background-color: #343B41;
                color: #FFFFFF;
                border: 1px solid #7289DA;
                border-radius: 10px;
                padding: 10px;
            }
            QComboBox QAbstractItemView {
                background-color: #343B41;
                color: #FFFFFF;
                border: 1px solid #7289DA;
                border-radius: 10px;
                padding: 10px;
            }
            QComboBox QAbstractItemView::item {
                background-color: #343B41;
                color: #FFFFFF;
                padding: 10px;
            }
            QComboBox QAbstractItemView::item:selected {
                background-color: #7289DA;
                color: #FFFFFF;
            }
            QComboBox QAbstractItemView::item:hover {
                background-color: #7289DA;
                color: #FFFFFF;
            }
        """)
        self.date_layout.addWidget(self.day_combobox)

        # Startadresse
        self.start_label = QLabel("Von:")
        self.start_label.setStyleSheet("color: #FFFFFF;")
        self.fahrten_layout.addWidget(self.start_label)
                
        self.start_combobox = QComboBox(self)
        self.start_combobox.setEditable(True)  # Damit die ComboBox editierbar ist
        self.start_combobox.setLineEdit(QLineEdit())  # Damit ein QLineEdit-Widget in der ComboBox verwendet wird
        self.start_combobox.addItems(self.get_streets())
        self.start_combobox.setStyleSheet("""
            QComboBox {
                background-color: #343B41;
                color: #FFFFFF;
                border: 1px solid #7289DA;
                border-radius: 10px;
                padding: 10px;
            }
            QComboBox QAbstractItemView {
                background-color: #343B41;
                color: #FFFFFF;
                border: 1px solid #7289DA;
                border-radius: 10px;
                padding: 10px;
            }
            QComboBox QAbstractItemView::item {
                background-color: #343B41;
                color: #FFFFFF;
                padding: 10px;
            }
            QComboBox QAbstractItemView::item:selected {
                background-color: #7289DA;
                color: #FFFFFF;
            }
            QComboBox QAbstractItemView::item:hover {
                background-color: #7289DA;
                color: #FFFFFF;
            }
        """)

        self.fahrten_layout.addWidget(self.start_combobox)

        # Zieladresse
        self.end_label = QLabel("Zu:")
        self.end_label.setStyleSheet("color: #FFFFFF;")
        self.fahrten_layout.addWidget(self.end_label)
                
        self.end_combobox = QComboBox(self)
        self.end_combobox.setEditable(True)  # Damit die ComboBox editierbar ist
        self.end_combobox.setLineEdit(QLineEdit())  # Damit ein QLineEdit-Widget in der ComboBox verwendet wird
        self.end_combobox.addItems(self.get_streets())
        self.end_combobox.setStyleSheet("""
            QComboBox {
                background-color: #343B41;
                color: #FFFFFF;
                border: 1px solid #7289DA;
                border-radius: 10px;
                padding: 10px;
            }
            QComboBox QAbstractItemView {
                background-color: #343B41;
                color: #FFFFFF;
                border: 1px solid #7289DA;
                border-radius: 10px;
                padding: 10px;
            }
            QComboBox QAbstractItemView::item {
                background-color: #343B41;
                color: #FFFFFF;
                padding: 10px;
            }
            QComboBox QAbstractItemView::item:selected {
                background-color: #7289DA;
                color: #FFFFFF;
            }
            QComboBox QAbstractItemView::item:hover {
                background-color: #7289DA;
                color: #FFFFFF;
            }
        """)

        self.fahrten_layout.addWidget(self.end_combobox)

        # Hinzufügen der Fahrt
        self.add_button = QPushButton("Hinzufügen", self)
        self.add_button.clicked.connect(self.add_fahrt)
        self.add_button.setStyleSheet("""QPushButton {
            background-color: #2C2F33;
            color: #FFFFFF;
            border-radius: 10px;
            padding: 10px;
            font-size: 16px;
            border: 1px solid #408df7;
        }
        QPushButton:hover {
            background-color: #408df7;
        }""")
        self.fahrten_layout.addWidget(self.add_button)

        # Liste der Fahrten
        self.fahrten_listbox = QListWidget(self)
        self.fahrten_listbox.setStyleSheet("""QListWidget {
            background-color: #343B41;
            color: #FFFFFF;
            border: 1px solid #7289DA;
            border-radius: 10px;
            padding: 10px;
        }""")
        self.fahrten_layout.addWidget(self.fahrten_listbox)

        # Speichern-Button
        self.save_button = QPushButton("Speichern", self)
        self.save_button.clicked.connect(self.save_to_excel)
        self.save_button.setStyleSheet("""QPushButton {
            background-color: #2C2F33;
            color: #FFFFFF;
            border-radius: 10px;
            padding: 10px;
            font-size: 16px;
            border: 1px solid #408df7;
        }
        QPushButton:hover {
            background-color: #408df7;
        }""")
        self.fahrten_layout.addWidget(self.save_button)

        self.stacked_widget.addWidget(self.fahrten_page)

        # Patienten-Seite
        self.patienten_page = QWidget()
        self.patienten_layout = QVBoxLayout(self.patienten_page)

        self.patienten_table_widget = QTableWidget()
        self.patienten_table_widget.setRowCount(len(patienten_daten))
        self.patienten_table_widget.setColumnCount(4)  # Eine Spalte mehr für die Checkbox
        self.patienten_table_widget.setHorizontalHeaderLabels(["Vorname", "Nachname", "Nummer", ""])  # Leere Spalte für die Checkbox

        # Anpassen der Tabelle an das Discord-ähnliche Design
        self.patienten_table_widget.setStyleSheet("""
            QTableWidget {
                background-color: #343B41;
                color: #FFFFFF;
                border: 1px solid #7289DA;
                border-radius: 10px;
                padding: 10px;
            }
            QTableWidget::item {
                background-color: #343B41;
                color: #FFFFFF;
                padding: 10px;
            }
            QTableWidget::item:selected {
                background-color: #7289DA;
                color: #FFFFFF;
            }
            QTableWidget::header {
                background-color: #2C2F33;
                color: #FFFFFF;
                padding: 10px;
            }
        """)

        # Anpassen der Spaltenbreite
        header = self.patienten_table_widget.horizontalHeader ()
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(2, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)  # Fixe Breite für die Checkbox-Spalte
        header.resizeSection(3, 50)  # Breite der Checkbox-Spalte auf 50 Pixel setzen

        # Abrunden der Ecken
        self.patienten_table_widget.setContentsMargins(10, 10, 10, 10)

        self.patienten_layout.addWidget(self.patienten_table_widget)

        self.add_patient_button = QPushButton("Patient hinzufügen")
        self.add_patient_button.clicked.connect(self.add_patient)
        self.add_patient_button.setStyleSheet("""QPushButton {
            background-color: #2C2F33;
            color: #FFFFFF;
            border-radius: 10px;
            padding: 10px;
            font-size: 16px;
            border: 1px solid #7289DA;
        }
        QPushButton:hover {
            background-color: #7289DA;
        }""")
        self.patienten_layout.addWidget(self.add_patient_button)

        self.delete_patient_button = QPushButton("Patient löschen")
        self.delete_patient_button.clicked.connect(self.delete_patient)
        self.delete_patient_button.setEnabled(False)
        self.delete_patient_button.setStyleSheet("""QPushButton {
            background-color: #2C2F33;
            color: #FFFFFF;
            border-radius: 10px;
            padding: 10 px;
            font-size: 16px;
            border: 1px solid #7289DA;
        }
        QPushButton:hover {
            background-color: #7289DA;
        }""")
        self.delete_patient_button.setStyleSheet("""
            QPushButton:disabled {
                background-color: #343B41;
                color: #FFFFFF;
                border-radius: 10px;
                padding: 10px;
                font-size: 16px;
                opacity: 0.5;
            }
            QPushButton:enabled {
                background-color: #2C2F33;
                color: #FFFFFF;
                border-radius: 10px;
                padding: 10px;
                font-size: 16px;
                border: 1px solid #FF0000;
            }
            QPushButton:hover:enabled {
                background-color: #7289DA;
                border: 1px solid #FF0000;
            }
        """)
        self.patienten_layout.addWidget(self.delete_patient_button)

        self.stacked_widget.addWidget(self.patienten_page)

        # Tabelle-Seite
        self.tabelle_page = QWidget()
        self.tabelle_layout = QVBoxLayout(self.tabelle_page)

        self.table_widget = QTableWidget()
        self.table_widget.setRowCount(len(strecken_daten))
        self.table_widget.setColumnCount(4)  # Eine Spalte mehr für die Checkbox
        self.table_widget.setHorizontalHeaderLabels(["Start", "Ziel", "Distanz (km)", ""])  # Leere Spalte für die Checkbox

        # Anpassen der Tabelle an das Discord-ähnliche Design
        self.table_widget.setStyleSheet("""
            QTableWidget {
                background-color: #343B41;
                color: #FFFFFF;
                border: 1px solid #7289DA;
                border-radius: 10px;
                padding: 10px;
            }
            QTableWidget::item {
                background-color: #343B41;
                color: #FFFFFF;
                padding: 10px;
            }
            QTableWidget::item:selected {
                background-color: #7289DA;
                color: #FFFFFF;
            }
            QTableWidget::header {
                background-color: #2C2F33;
                color: #FFFFFF;
                padding: 10px;
            }
        """)

        # Anpassen der Spaltenbreite
        header = self.table_widget.horizontalHeader()
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(2, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)  # Fixe Breite für die Checkbox-Spalte
        header.resizeSection(3, 50)  # Breite der Checkbox-Spalte auf 50 Pixel setzen

        # Abrunden der Ecken
        self.table_widget.setContentsMargins(10, 10, 10, 10)

        self.tabelle_layout.addWidget(self.table_widget)

        self.add_strecke_button = QPushButton("Neue Strecke hinzufügen")
        self.add_strecke_button.clicked.connect(self.add_strecke)
        self.add_strecke_button.setStyleSheet("""QPushButton {
            background-color: #2C2F33;
            color: #FFFFFF;
            border-radius: 10px;
            padding: 10px;
            font-size: 16px;
            border: 1px solid #7289DA;
        }
        QPushButton:hover {
            background-color: #7289DA;
        }""")
        self.tabelle_layout.addWidget(self.add_strecke_button)

        self.delete_strecke_button = QPushButton("Strecke löschen")
        self.delete_strecke_button.clicked.connect(self.delete_strecke)
        self.delete_strecke_button.setEnabled(False)
        self.delete_strecke_button.setStyleSheet("""QPushButton {
            background-color: #2C2F33;
            color: #FFFFFF;
            border-radius: 10px;
            padding: 10px;
            font-size: 16px;
            border: 1px solid #7289DA;
        }
        QPushButton:hover {
            background-color: #7289DA;
        }""")
        self.delete_strecke_button.setStyleSheet("""
            QPushButton:disabled {
                background-color: #343B41;
                color: #FFFFFF;
                border-radius: 10px;
                padding: 10px;
                font-size: 16px;
                opacity: 0.5;
            }
            QPushButton:enabled {
                background-color: #2C2F33;
                color: #FFFFFF;
                border-radius: 10px;
                padding: 10px;
                font-size: 16px;
                border: 1px solid #FF0000;
            }
            QPushButton:hover:enabled {
                background-color: #7289DA;
                border: 1px solid #FF0000;
            }
        """)
        
        self.tabelle_layout.addWidget(self.delete_strecke_button)

        self.stacked_widget.addWidget(self.tabelle_page)

        self.patienten_table_widget.itemChanged.connect(self.patient_table_item_changed)
        self.table_widget.itemChanged.connect(self.table_item_changed)

        # Sonstiges Seite erstellen
        self.sonstiges_page = QWidget()
        self.sonstiges_layout = QVBoxLayout(self.sonstiges_page)
        self.stacked_widget.addWidget(self.sonstiges_page)

        # Layouts kombinieren
        main_layout.addWidget(menu_frame)
        main_layout.addWidget(self.stacked_widget)

        self.update_patient_table()  # Aktualisiere die Tabelle mit den geladenen Daten

    def show_fahrten(self):
        self.stacked_widget.setCurrentIndex(0)  # Set the current index to the fahrten page
        logger.info('Hauptmenüindex gesetzt')

    def show_tabelle(self):
        self.stacked_widget.setCurrentIndex(1)  # Set the current index to the tabelle page
        logger.info('Streckenindex gesetzt')

    def show_patienten(self):
        self.stacked_widget.setCurrentIndex(2)  # Set the current index to the patienten page
        logger.info('Patientenindex gesetzt')

    def exportieren(self):
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle('Excel-Tabelle exportieren')
        msg_box.setText('Sind Sie sicher, dass Sie die aktuelle Excel-Tabelle mit allen Einträgen exportieren möchten?')
        msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        
        # Ändern der Textfarbe auf Weiß
        msg_box.setStyleSheet("color: #FFFFFF; background-color: #2C2F33;")
        
        # Ändern des Styles der Buttons
        yes_button = msg_box.button(QMessageBox.Yes)
        no_button = msg_box.button(QMessageBox.No)
        
        yes_button.setText('Ja')
        no_button.setText('Nein')
        
        yes_button.setStyleSheet("""
            QPushButton {
                background-color: #2C2F33;
                color: #FFFFFF;
                border-radius: 10px;
                padding: 10px;
                font-size: 16px;
                border: 1px solid #7289DA;
            }
            QPushButton:hover {
                background-color: #7289DA;
            }
        """)
        
        no_button.setStyleSheet("""
            QPushButton {
                background-color: #2C2F33;
                color: #FFFFFF;
                border-radius: 10px;
                padding: 10px;
                font-size: 16px;
                border: 1px solid #7289DA;
            }
            QPushButton:hover {
                background-color: #7289DA;
            }
        """)
        
        reply = msg_box.exec_()
        
        if reply == QMessageBox.Yes:
            # Create a file dialog to select the save location and file name
            file_dialog = QtW.QFileDialog(self)
            file_dialog.setFileMode(QtW.QFileDialog.AnyFile)
            file_dialog.setAcceptMode(QtW.QFileDialog.AcceptSave)
            file_dialog.setDirectory(str(base_dir / "Excel"))
            file_dialog.setNameFilters(["Excel files (*.xlsx)"])
            file_dialog.selectFile("fahrtenbuch_export.xlsx")
            selected_file_path = file_dialog.selectedFiles()[0]
            
            if file_dialog.exec_():
                file_name = file_dialog.selectedFiles()[0]
                # Copy the existing Excel file to the selected location
                shutil.copy2(str(EXCEL_FILE), file_name)
                msg_box = QMessageBox(QMessageBox.Information, "Erfolgreich", "Die Tabelle wurde erfolgreich exportiert.")
                style_sheet = """
                QMessageBox {
                    background-color: #2C2F33;
                    color: #FFFFFF;
                    border-radius: 10px;
                    padding: 10px;
                    font-size: 16px;
                    border: 1px solid #7289DA;
                }
                QMessageBox QPushButton {
                    background-color: #2C2F33;
                    color: #FFFFFF;
                    border-radius: 10px;
                    padding: 10px;
                    font-size: 16px;
                    border: 1px solid #7289DA;
                }
                QMessageBox QLabel {
                    color: #FFFFFF;
                }
                QMessageBox QPushButton:hover {
                    background-color: #7289DA;
                }
                """
                msg_box.setStyleSheet(style_sheet)
                msg_box.exec_()
                logger.info('Tabelle wurde exportiert nach {selected_file_path}')
        else:
            logger.info('Datei wurde nicht exportiert')

    def check_version(self):
        with open(VERSION_FILE, 'r') as f:
            local_version_data = json.load(f)
        local_version = local_version_data['version']

        print("local_version:", local_version)

        response = requests.get(GITHUB_REPO_URL)

        if response.status_code == 200:
            github_version_data = response.json()
            github_version = github_version_data['version']
            print("github_version:", github_version)

            if github_version and github_version > local_version:
                msg_box = QMessageBox(self)
                msg_box.setWindowTitle('Update Gefunden')
                msg_box.setText('Es wurde ein Update für die App gefunden.\nWollen Sie dieses installieren?')
                msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
                
                # Ändern der Textfarbe auf Weiß
                msg_box.setStyleSheet("color: #FFFFFF; background-color: #2C2F33;")
                
                # Ändern des Styles der Buttons
                yes_button = msg_box.button(QMessageBox.Yes)
                no_button = msg_box.button(QMessageBox.No)
                
                yes_button.setText('Ja')
                no_button.setText('Nein')
                
                yes_button.setStyleSheet("""
                    QPushButton {
                        background-color: #2C2F33;
                        color: #FFFFFF;
                        border-radius: 10px;
                        padding: 10px;
                        font-size: 16px;
                        border: 1px solid #7289DA;
                    }
                    QPushButton:hover {
                        background-color: #7289DA;
                    }
                """)
                
                no_button.setStyleSheet("""
                    QPushButton {
                        background-color: #2C2F33;
                        color: #FFFFFF;
                        border-radius: 10px;
                        padding: 10px;
                        font-size: 16px;
                        border: 1px solid #7289DA;
                    }
                    QPushButton:hover {
                        background-color: #7289DA;
                    }
                """)
                
                reply = msg_box.exec_()
                if reply == QMessageBox.Yes:
                    msg_box = QMessageBox(QMessageBox.Information, "Allesklar!", "Die App wird nun geschlossen.\nGleich geht es weiter.")
                    style_sheet = """
                    QMessageBox {
                        background-color: #2C2F33;
                        color: #FFFFFF;
                        border-radius: 10px;
                        padding: 10px;
                        font-size: 16px;
                        border: 1px solid #7289DA;
                    }
                    QMessageBox QPushButton {
                        background-color: #2C2F33;
                        color: #FFFFFF;
                        border-radius: 10px;
                        padding: 10px;
                        font-size: 16px;
                        border: 1px solid #7289DA;
                    }
                    QMessageBox QLabel {
                        color: #FFFFFF;
                    }
                    QMessageBox QPushButton:hover {
                        background-color: #7289DA;
                    }
                    """
                    msg_box.setStyleSheet(style_sheet)
                    msg_box.accepted.connect(self.on_accepted)
                    msg_box.exec_()

            else:
                msg_box = QMessageBox(QMessageBox.Information, "Kein Update verfügbar", "Kein Update gefunden.\nSie sind auf dem neusten Stand.")
                style_sheet = """
                QMessageBox {
                    background-color: #2C2F33;
                    color: #FFFFFF;
                    border-radius: 10px;
                    padding: 10px;
                    font-size: 16px;
                    border: 1px solid #7289DA;
                }
                QMessageBox QPushButton {
                    background-color: #2C2F33;
                    color: #FFFFFF;
                    border-radius: 10px;
                    padding: 10px;
                    font-size: 16px;
                    border: 1px solid #7289DA;
                }
                QMessageBox QLabel {
                    color: #FFFFFF;
                }
                QMessageBox QPushButton:hover {
                    background-color: #7289DA;
                }
                """
                msg_box.setStyleSheet(style_sheet)
                msg_box.exec_()
        else:
            print('Error reading GitHub version file')

    def update_patient_combobox(self):
        """Aktualisiere die Patientenauswahl in der ComboBox."""
        self.patient_combobox.clear()
        for patient in patienten_daten:
            self.patient_combobox.addItem(f"{patient['vorname']} {patient['nachname']} ({patient['nummer']})")
            logger.info('Patientencombobox wurde geupdated')

    def add_patient(self):
        """Füge einen neuen Patienten hinzu."""
        vorname, ok_vorname = QInputDialog.getText(self, 'Neuer Patient', 'Vorname:')
        nachname, ok_nachname = QInputDialog.getText(self, 'Neuer Patient', 'Nachname:')
        nummer, ok_nummer = QInputDialog.getText(self, 'Neuer Patient', 'Nummer:')
        
        if ok_vorname and ok_nachname and ok_nummer:
            patienten_daten.append({"vorname": vorname, "nachname": nachname, "nummer": nummer})
            save_patienten_daten(patienten_daten)
            self.update_patient_table()
            self.update_patient_combobox()
            logger.info('Patient wurde hinzugefügt')

    def delete_patient(self):
        """Lösche ausgewählten Patienten."""
        checked_rows = []
        for row in range(self.patienten_table_widget.rowCount()):
            checkbox = self.patienten_table_widget.item(row, 3)
            if checkbox is not None and checkbox.checkState() == Qt.Checked:
                checked_rows.append(row)
        
        if checked_rows:
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle('Patient löschen')
            msg_box.setText('Sind Sie sicher, dass Sie diesen Patienten löschen möchten?')
            msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            
            # Ändern der Textfarbe auf Weiß
            msg_box.setStyleSheet("color: #FFFFFF; background-color: #2C2F33;")
            
            # Ändern des Styles der Buttons
            yes_button = msg_box.button(QMessageBox.Yes)
            no_button = msg_box.button(QMessageBox.No)
            
            yes_button.setText('Ja')
            no_button.setText('Nein')
            
            yes_button.setStyleSheet("""
                QPushButton {
                    background-color: #2C2F33;
                    color: #FFFFFF;
                    border-radius: 10px;
                    padding: 10px;
                    font-size: 16px;
                    border: 1px solid #7289DA;
                }
                QPushButton:hover {
                    background-color: #7289DA;
                }
            """)
            
            no_button.setStyleSheet("""
                QPushButton {
                    background-color: #2C2F33;
                    color: #FFFFFF;
                    border-radius: 10px;
                    padding: 10px;
                    font-size: 16px;
                    border: 1px solid #7289DA;
                }
                QPushButton:hover {
                    background-color: #7289DA;
                }
            """)
            
            reply = msg_box.exec_()
            if reply == QMessageBox.Yes:
                rows = sorted(checked_rows, reverse=True)
                for row in rows:
                    del patienten_daten[row]
                save_patienten_daten(patienten_daten)
                self.update_patient_table()
                logger.info('Patient wurde gelöscht')


    def on_accepted(self):
      self.close()
      QApplication.quit()  # Beendet die PyQt-Anwendung vollständig
      os.system(f"python {str(UPDATE_FILE)}")
      sys.exit()  # Beendet den gesamten Python-Prozess

    def update_patient_table(self):
        """Aktualisiere die Tabelle der Patienten."""
        self.patienten_table_widget.setRowCount(len(patienten_daten))
        for row, patient in enumerate(patienten_daten):
            self.patienten_table_widget.setItem(row, 0, QTableWidgetItem(patient['vorname']))
            self.patienten_table_widget.setItem(row, 1, QTableWidgetItem(patient['nachname']))
            self.patienten_table_widget.setItem(row, 2, QTableWidgetItem(patient['nummer']))
            checkbox = QTableWidgetItem()
            checkbox.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            checkbox.setCheckState(Qt.Unchecked)
            self.patienten_table_widget.setItem(row, 3, checkbox)
        self.delete_patient_button.setEnabled(False)
        logger.info('Patiententabelle wurde geupdated')

    def patient_table_item_changed(self, item):
        """Aktiviert oder deaktiviert den delete_patient_button, je nachdem, ob eine Zeile ausgewählt ist."""
        if item.column() == 3:  # Checkbox-Spalte
            checked_rows = 0
            for row in range(self.patienten_table_widget.rowCount()):
                checkbox = self.patienten_table_widget.item(row, 3)
                if checkbox is not None and checkbox.checkState() == Qt.Checked:
                    checked_rows += 1
            self.delete_patient_button.setEnabled(checked_rows > 0)
            logger.info('Patientenbutton aktiviert')
        else:
            logger.info('Patientenbutton deaktiviert')

    def table_item_changed(self, item):
        """Aktiviert oder deaktiviert den delete_strecke_button, je nachdem, ob eine Zeile ausgewählt ist."""
        if item.column() == 3:  # Checkbox-Spalte
            checked_rows = 0
            for row in range(self.table_widget.rowCount()):
                checkbox = self.table_widget.item(row, 3)
                if checkbox is not None and checkbox.checkState() == Qt.Checked:
                    checked_rows += 1
            self.delete_strecke_button.setEnabled(checked_rows > 0)
            logger.info('Streckenbutton aktiviert')

        else:
            logger.info('Streckenbutton deaktiviert')

    def add_fahrt(self):
        """Fügt eine Fahrt zur Liste der Fahrten hinzu."""
        year = self.year_combobox.currentText()
        month = self.month_combobox.currentText()
        day = self.day_combobox.currentText()
        date = f"{year}-{month}-{day}"
        start = self.start_combobox.currentText()
        ziel = self.end_combobox.currentText()
        patient = self.patient_combobox.currentText()

        distanz = next((strecke['distanz'] for strecke in strecken_daten 
                        if strecke['start'] == start and strecke['ziel'] == ziel), None)

        if start and ziel and date and distanz is not None and patient:
            self.fahrten_listbox.addItem(f"{date}: {patient} - {start} -> {ziel} ({distanz} km)")
            self.fahrten.append({
                'datum': date,
                'start': start,
                'ziel': ziel,
                'distanz': distanz,
                'patient': patient
            })
            logger.info('Fahrt wurde hinzugefügt')
        else:
            msg_box = QMessageBox(QMessageBox.Warning, "Fehler", "Wählen Sie eine gültige Strecke und Patienten.")
            style_sheet = """
            QMessageBox {
                background-color: #2C2F33;
                color: #FFFFFF;
                border-radius: 10px;
                padding: 10px;
                font-size: 16px;
                border: 1px solid #7289DA;
            }
            QMessageBox QPushButton {
                background-color: #2C2F33;
                color: #FFFFFF;
                border-radius: 10px;
                padding: 10px;
                font-size: 16px;
                border: 1px solid #7289DA;
            }
            QMessageBox QLabel {
                color: #FFFFFF;
            }
            QMessageBox QPushButton:hover {
                background-color: #7289DA;
            }
            """
            msg_box.setStyleSheet(style_sheet)
            msg_box.exec_()
            logger.warning('Fahrt konnte nicht hinzugefügt werden')
            logger.warning('Keine Strecke/Patient gewählt')

    def save_to_excel(self):
        """Speichere die Fahrten in eine Excel-Datei."""
        if not self.fahrten:  # Check if the list of trips is empty
            msg_box = QMessageBox(QMessageBox.Information, "Keine Fahrten gewählt", "Sie haben keine Fahrten gewählt.")
            style_sheet = """
            QMessageBox {
                background-color: #2C2F33;
                color: #FFFFFF;
                border-radius: 10px;
                padding: 10px;
                font-size: 16px;
                border: 1px solid #7289DA;
            }
            QMessageBox QPushButton {
                background-color: #2C2F33;
                color: #FFFFFF;
                border-radius: 10px;
                padding: 10px;
                font-size: 16px;
                border: 1px solid #7289DA;
            }
            QMessageBox QPushButton:hover {
                background-color: #7289DA;
            }
            """
            msg_box.setStyleSheet(style_sheet)
            msg_box.exec_()
            logger.warning('Keine Fahrten wurden gefunden | Listbox empty')
            return  # Exit the function early

        if os.path.exists(EXCEL_FILE):
            try:
                os.rename(EXCEL_FILE, EXCEL_FILE)
            except OSError as e:
                msg_box = QMessageBox(QMessageBox.Information, "Fehler - Datei Geöffnet", f"Die Datei läuft bereits im Hintergrund, schließen Sie diese.")
                style_sheet = """
                QMessageBox {
                    background-color: #2C2F33;
                    color: #FFFFFF;
                    border-radius: 10px;
                    padding: 10px;
                    font-size: 16px;
                }
                QMessageBox QLabel {
                    color: #FFFFFF;
                }
                QMessageBox QPushButton {
                    background-color: #2C2F33;
                    color: #FFFFFF;
                    border-radius: 10px;
                    padding: 10px;
                    font-size: 16px;
                    border: 1px solid #7289DA;
                }
                QMessageBox QPushButton:hover {
                    background-color: #7289DA;
                }
                """
                msg_box.setStyleSheet(style_sheet)
                msg_box.exec_()
                logger.warning('Datei bereits geöffnet | Conflict')
                return

        if os.path.exists(EXCEL_FILE):
            workbook = load_workbook(EXCEL_FILE)
            sheet = workbook.active
            logger.info('Datei gefunden')
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Fahrten"
            # Header für die Excel-Datei
            sheet.append(["Patient", "Datum", "Von", "Zu", "Distanz (km)"])
            logger.info('Datei nicht gefunden | Datei wurde erstellt')

        # Berechne die Summe der bestehenden Kilometer in der Tabelle (Spalte Distanz)
        gesamt_kilometer = 0
        for row in range(2, sheet.max_row + 1):  # Beginne bei Zeile 2, um die Kopfzeile zu überspringen
            distanz = sheet.cell(row=row, column=5).value  # Spalte 5 ist die Distanz (km)
            if isinstance(distanz, (int, float)):  # Überprüfe, ob es sich um eine Zahl handelt
                gesamt_kilometer += distanz
                logger.info('Berechnung gesamtkilometer ({gesamt_kilometer})')

        # Füge die neuen Fahrten hinzu und aktualisiere die Kilometer
        for fahrt in self.fahrten:
            sheet.append([fahrt['patient'], fahrt['datum'], fahrt['start'], fahrt['ziel'], fahrt['distanz']])
            gesamt_kilometer += fahrt['distanz']
            logger.info('Fahrt hinzugefügt')

        # Suche nach der Zeile, die bereits "Jahreskilometer:" enthält, falls vorhanden, und lösche sie
        jahreskilometer_row = None
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == "Jahreskilometer:":
                jahreskilometer_row = row
                logger.info('Jahreskilometerzeile gefunden')
                break

        if jahreskilometer_row is not None:
            sheet.delete_rows(jahreskilometer_row)
            logger.info('Jahreskilometerzeile löschen')

        # Füge die Zeile mit den Jahreskilometern am Ende hinzu
        sheet.append(["Jahreskilometer:", gesamt_kilometer])
        logger.info('Jahreskilometerzeile neu hinzufügen')

        # Speichere die Excel-Datei
        workbook.save(EXCEL_FILE)
        logger.info('Datei aktualisiert')

        # Erzeuge eine Messagebox mit einer Erfolgsmeldung
        msg_box = QMessageBox(QMessageBox.Information, "Gespeichert", f"Fahrten erfolgreich in '{EXCEL_FILE}' gespeichert.")
        style_sheet = """
        QMessageBox {
            background-color: #2C2F33;
            color: #FFFFFF;
            border-radius: 10px;
            padding: 10px;
            font-size: 16px;
            border: 1px solid #7289 DA;
        }
        QMessageBox QPushButton {
            background-color: #2C2F33;
            color: #FFFFFF;
            border-radius: 10px;
            padding: 10px;
            font-size: 16px;
            border: 1px solid #7289DA;
        }
        QMessageBox QPushButton:hover {
            background-color: #7289DA;
        }
        """
        msg_box.setStyleSheet(style_sheet)
        msg_box.exec_()

        # Leere die Box über dem Button
        self.fahrten_listbox.clear()
        logger.info('Fahrtenbox gelöscht')


    def get_streets(self):
        """Lade die Straßen aus den Streckendaten."""
        streets = set()
        for street in strecken_daten:
            streets.add(street['start'])
            streets.add(street['ziel'])
            logger.info('Streckendaten laden')
        return list(streets)

    def add_strecke(self):
        """Füge eine neue Strecke hinzu."""
        von, ok_von = QInputDialog.getText(self, 'Neue Strecke', 'Von:')
        zu, ok_zu = QInputDialog.getText(self, 'Neue Strecke', 'Zu:')
        distanz, ok_distanz = QInputDialog.getDouble(self, 'Neue Strecke', 'Distanz (km):')

        if ok_von and ok_zu and ok_distanz >= 0:
            strecken_daten.append({"start": von, "ziel": zu, "distanz": distanz})
            save_strecken_daten(strecken_daten)
            self.update_table()
            logger.info('Strecke hinzugefügt')

    def delete_strecke(self):
        """Lösche ausgewählte Strecken."""
        checked_rows = []
        for row in range(self.table_widget.rowCount()):
            checkbox = self.table_widget.item(row, 3)
            if checkbox.checkState() == Qt.Checked:
                checked_rows.append(row)
        
        if checked_rows:
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle('Strecken löschen')
            msg_box.setText('Sind Sie sicher, dass Sie diese Strecken löschen möchten?')
            msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            
            # Ändern der Textfarbe auf Weiß
            msg_box.setStyleSheet("color: #FFFFFF; background-color: #2C2F33;")
            
            # Ändern des Styles der Buttons
            yes_button = msg_box.button(QMessageBox.Yes)
            no_button = msg_box.button(QMessageBox.No)
            
            yes_button.setText('Ja')
            no_button.setText('Nein')
            
            yes_button.setStyleSheet("""
                QPushButton {
                    background-color: #2C2F33;
                    color: #FFFFFF;
                    border-radius: 10px;
                    padding: 10px;
                    font-size: 16px;
                    border: 1px solid #7289DA;
                }
                QPushButton:hover {
                    background-color: #7289DA;
                }
            """)
            
            no_button.setStyleSheet("""
                QPushButton {
                    background-color: #2C2F33;
                    color: #FFFFFF;
                    border-radius: 10px;
                    padding: 10px;
                    font-size: 16px;
                    border: 1px solid #7289DA;
                }
                QPushButton:hover {
                    background-color: #7289DA;
                }
            """)
            
            reply = msg_box.exec_()
            if reply == QMessageBox.Yes:
                rows = sorted(checked_rows, reverse=True)
                for row in rows:
                    del strecken_daten[row]
                save_strecken_daten(strecken_daten)
                self.update_table()
                logger.info('Strecke ({strecken_daten[row]}) gelöscht')

    def update_table(self):
        """Aktualisiere die Tabelle mit den Streckendaten."""
        self.table_widget.setRowCount(len(strecken_daten))
        for row, strecke in enumerate(strecken_daten):
            self.table_widget.setItem(row, 0, QTableWidgetItem(strecke['start']))
            self.table_widget.setItem(row, 1, QTableWidgetItem(strecke['ziel']))
            self.table_widget.setItem(row, 2, QTableWidgetItem(str(strecke['distanz'])))
            checkbox = QTableWidgetItem()
            checkbox.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            checkbox.setCheckState(Qt.Unchecked)
            self.table_widget.setItem(row, 3, checkbox)
        self.delete_strecke_button.setEnabled(False)

    def show_fahrten(self):
        self.stacked_widget.setCurrentIndex(0)
        logger.info('Wechsel zu Hauptmenüseite')

    def show_tabelle(self):
        """Wechselt zur Tabelle-Seite."""
        self.update_table()  # Aktualisiere die Tabelle vor dem Anzeigen
        self.stacked_widget.setCurrentIndex(2)  # Set the current index to the tabelle page
        logger.info('Wechsel zu Streckenseite')

    def show_patienten(self):
        """Wechselt zur Patientenverwaltung-Seite."""
        self.update_patient_table()  # Aktualisiere die Tabelle vor dem Anzeigen
        self.stacked_widget.setCurrentIndex(1)  # Set the current index to the patienten page
        logger.info('Wechsel zu Patientenseite')

    def close_app(self):
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle('App Schließen')
        msg_box.setText('Sind Sie sicher, dass Sie die App schließen möchten?')
        msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        
        # Ändern der Textfarbe auf Weiß
        msg_box.setStyleSheet("color: #FFFFFF; background-color: #2C2F33;")
        
        # Ändern des Styles der Buttons
        yes_button = msg_box.button(QMessageBox.Yes)
        no_button = msg_box.button(QMessageBox.No)
        
        yes_button.setText('Ja')
        no_button.setText('Nein')
        
        yes_button.setStyleSheet("""
            QPushButton {
                background-color: #2C2F33;
                color: #FFFFFF;
                border-radius: 10px;
                padding: 10px;
                font-size: 16px;
                border: 1px solid #7289DA;
            }
            QPushButton:hover {
                background-color: #7289DA;
            }
        """)
        
        no_button.setStyleSheet("""
            QPushButton {
                background-color: #2C2F33;
                color: #FFFFFF;
                border-radius: 10px;
                padding: 10px;
                font-size: 16px;
                border: 1px solid #7289DA;
            }
            QPushButton:hover {
                background-color: #7289DA;
            }
        """)
        
        reply = msg_box.exec_()
        if reply == QMessageBox.Yes:
            logger.info('App geschlossen')
            self.close()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = CustomWindow()
    window.show()
    sys.exit(app.exec_())


# test if it worked
